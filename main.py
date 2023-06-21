import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from tqdm import tqdm

# Load the XLSX file
xlsx_file = 'list.xlsx'
wb = openpyxl.load_workbook(xlsx_file)
sheet = wb.active

# Create a new workbook to store the extracted data
output_file = 'output_file.xlsx'
output_wb = Workbook()
output_sheet = output_wb.active

# Add headers to the output sheet
output_sheet.append(['Title', 'Description', 'H1', 'H2', 'H3', 'H4', 'H5', 'H6'])

# Get the total number of URLs
total_urls = sheet.max_row - 1

# Iterate through the URLs in the first column with a progress bar
for row in tqdm(sheet.iter_rows(min_row=2, values_only=True), total=total_urls, desc="Processing URLs"):
    url = row[0]

    # Send a request to the URL with the provided user agent and parse the HTML content
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 YaBrowser/23.5.2.625 Yowser/2.5 Safari/537.36'}
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Extract the required information
    title = soup.title.string if soup.title else ''
    description = soup.find('meta', {'name': 'description'})['content'] if soup.find('meta', {'name': 'description'}) else ''

    h_tags = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
    headers = [h_tag.text.strip() for h_tag in h_tags]

    # Add the extracted data to the output sheet
    output_sheet.append([title, description] + headers)

# Save the output workbook
output_wb.save(output_file)

print('Successful completion! Extracted data has been saved to', output_file)
